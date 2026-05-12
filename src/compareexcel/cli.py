"""Command-line interface for comparing two Excel files."""

from __future__ import annotations

import argparse
import os
import sys

import pandas as pd
from openpyxl import load_workbook

from compareexcel.core import (
    ALIGN_ONLY_NO_DATA_NOTE,
    compare_cell_formatting_sampled,
    compare_currency_column_totals,
    compare_data_alignment,
    compare_data_cells_alignment_and_format,
    compare_formatting,
    compare_header_alignment,
    compare_numeric_column_totals,
    compare_numeric_column_totals_quick,
    sheet_data_presence_mismatch,
    sheet_row_count_rows,
)
from compareexcel.report import write_report


def _ensure_utf8_stdio():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def _print_section(title: str, rows: list, describe_empty: str):
    print(title)
    if not rows:
        print(f"  {describe_empty}")
        return
    for r in rows:
        if isinstance(r, dict):
            parts = [f"{k}: {v}" for k, v in r.items() if v not in ("", None)]
            line = "  " + " | ".join(parts)
            if str(r.get("Note", "")).strip().lower() == ALIGN_ONLY_NO_DATA_NOTE.lower():
                line = f"\033[32m{line}\033[0m"
            print(line)
        else:
            print(f"  {r}")
    print()


def _print_sheet_name_comparison(path1: str, path2: str, wb1, wb2) -> None:
    """Print every sheet from each file with match / missing-from-other-file labels."""
    names1 = list(wb1.sheetnames)
    names2 = list(wb2.sheetnames)
    set1 = set(names1)
    set2 = set(names2)
    base1 = os.path.basename(path1)
    base2 = os.path.basename(path2)

    print()
    print("=== Sheet name comparison ===")
    print(f"File 1: {base1}")
    print(f"File 2: {base2}")
    print()

    print(f"All sheets in File 1 ({len(names1)}), in workbook order:")
    for name in names1:
        if name in set2:
            print(f"  [MATCH]     {name!r}  — present in File 2")
        else:
            print(f"  [MISSING]   {name!r}  — not in File 2 (only in File 1)")
    print()

    print(f"All sheets in File 2 ({len(names2)}), in workbook order:")
    for name in names2:
        if name in set1:
            print(f"  [MATCH]     {name!r}  — present in File 1")
        else:
            print(f"  [MISSING]   {name!r}  — not in File 1 (only in File 2)")
    print()

    common = sorted(set1 & set2)
    only1 = sorted(set1 - set2)
    only2 = sorted(set2 - set1)
    print("Summary — sheet names:")
    print(f"  Matching (in both files): {len(common)}")
    if common:
        for n in common:
            print(f"    {n!r}")
    else:
        print("    (none)")
    print(f"  Only in File 1: {len(only1)}")
    if only1:
        for n in only1:
            print(f"    {n!r}")
    else:
        print("    (none)")
    print(f"  Only in File 2: {len(only2)}")
    if only2:
        for n in only2:
            print(f"    {n!r}")
    else:
        print("    (none)")
    print()


def main():
    _ensure_utf8_stdio()

    parser = argparse.ArgumentParser(
        description=(
            "Compare two Excel workbooks: formatting, currency totals, header alignment, "
            "and data-cell alignment (full mode), or alignment + number format on data rows only "
            "(--alignment-only)."
        ),
    )
    parser.add_argument("file1", help="First workbook (.xlsx)")
    parser.add_argument("file2", help="Second workbook (.xlsx)")
    parser.add_argument("--sheet", help="Compare only this sheet (must exist in both)")
    parser.add_argument(
        "--alignment-only",
        "--ao",
        action="store_true",
        help=(
            "Header alignment and one compared cell per column (first row where both files have "
            "data in that column), including alignment and number_format. Columns with no such row "
            "are listed as 'no data' (green in report). Skips pandas totals, column summary, and "
            "sampled format pass. --sample-fraction is ignored."
        ),
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Write report to this path (.xlsx or .html)",
    )
    parser.add_argument(
        "--sample-fraction",
        type=float,
        default=0.1,
        help=(
            "Full mode only: fraction of per-column rows (both sides non-blank) to check for "
            "cell number_format (default: 0.1). Ignored with --alignment-only."
        ),
    )
    parser.add_argument(
        "--numeric-column-totals",
        "-nct",
        action="store_true",
        help=(
            "Full mode only: quick per-column sums using pandas only (no Excel number_format "
            "checks). Prints a console section; with --output, adds sheet Quick_Numeric_Column_Totals "
            "(and matching HTML section). Ignored with --alignment-only."
        ),
    )
    parser.add_argument(
        "--compare-sheet-names",
        action="store_true",
        help=(
            "List every sheet in File 1 and File 2 (workbook order), label each as matching the "
            "other file or missing from it, print a summary of names only in one file, then exit. "
            "Does not run formatting, alignment, or totals comparisons."
        ),
    )
    parser.add_argument(
        "--row-counts",
        action="store_true",
        help=(
            "Print a per-sheet row-count summary (openpyxl Worksheet.max_row) after the run "
            "header, and add a Rowcounts sheet (Excel) or HTML section when using --output. "
            "With --sheet, only that sheet is listed; otherwise every sheet name in either file."
        ),
    )

    args = parser.parse_args()

    if not os.path.isfile(args.file1):
        print(f"Error: file not found: {args.file1}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(args.file2):
        print(f"Error: file not found: {args.file2}", file=sys.stderr)
        sys.exit(1)

    if not (0 < args.sample_fraction <= 1):
        print("--sample-fraction must be between 0 and 1 (exclusive of 0).", file=sys.stderr)
        sys.exit(1)

    wb1 = load_workbook(args.file1, data_only=False)
    wb2 = load_workbook(args.file2, data_only=False)

    if args.compare_sheet_names:
        _print_sheet_name_comparison(args.file1, args.file2, wb1, wb2)
        sys.exit(0)

    names1 = set(wb1.sheetnames)
    names2 = set(wb2.sheetnames)
    if args.sheet:
        if args.sheet not in names1 or args.sheet not in names2:
            print(
                f"Error: sheet {args.sheet!r} missing from one or both workbooks.",
                file=sys.stderr,
            )
            sys.exit(1)
        sheets = [args.sheet]
    else:
        sheets = sorted(names1 & names2)
        if names1 != names2:
            print("Sheet names differ between files.")
            if names1 - names2:
                print(f"  Only in file 1: {sorted(names1 - names2)}")
            if names2 - names1:
                print(f"  Only in file 2: {sorted(names2 - names1)}")
            print(f"  Comparing common sheets only: {sheets}\n")

    row_count_rows: list | None = None
    if args.row_counts:
        row_count_rows = sheet_row_count_rows(
            wb1, wb2, sheets=sheets if args.sheet else None
        )

    dfs1: dict = {}
    dfs2: dict = {}
    if not args.alignment_only:
        try:
            if args.sheet:
                dfs1 = {args.sheet: pd.read_excel(args.file1, sheet_name=args.sheet)}
                dfs2 = {args.sheet: pd.read_excel(args.file2, sheet_name=args.sheet)}
            else:
                dfs1 = pd.read_excel(args.file1, sheet_name=None)
                dfs2 = pd.read_excel(args.file2, sheet_name=None)
        except Exception as e:
            print(f"Error reading workbooks with pandas: {e}", file=sys.stderr)
            sys.exit(1)

    all_column_fmt: list[dict] = []
    all_cell_fmt: list[dict] = []
    all_currency_report: list[dict] = []
    all_numeric_report: list[dict] = []
    all_quick_numeric: list[dict] = []
    all_header_align: list[dict] = []
    all_data_align: list[dict] = []
    all_align_and_format: list[dict] = []
    all_sheet_presence: list[dict] = []

    for sheet in sheets:
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]

        if not args.alignment_only:
            if sheet not in dfs1 or sheet not in dfs2:
                continue

        pres = sheet_data_presence_mismatch(ws1, ws2, sheet)
        if pres:
            all_sheet_presence.append(pres)

        if not args.alignment_only:
            df1 = dfs1[sheet]
            df2 = dfs2[sheet]
            for row in compare_formatting(ws1, ws2):
                row = {**row, "Sheet": sheet}
                all_column_fmt.append(row)
            for row in compare_currency_column_totals(df1, df2, ws1, ws2):
                row = {**row, "Sheet": sheet}
                all_currency_report.append(row)
            for row in compare_numeric_column_totals(df1, df2, ws1, ws2):
                row = {**row, "Sheet": sheet}
                all_numeric_report.append(row)
            if args.numeric_column_totals:
                for row in compare_numeric_column_totals_quick(df1, df2):
                    row = {**row, "Sheet": sheet}
                    all_quick_numeric.append(row)
            for row in compare_cell_formatting_sampled(ws1, ws2, args.sample_fraction):
                row = {**row, "Sheet": sheet}
                all_cell_fmt.append(row)
            for row in compare_data_alignment(ws1, ws2):
                row = {**row, "Sheet": sheet}
                all_data_align.append(row)

        for row in compare_header_alignment(ws1, ws2):
            row = {**row, "Sheet": sheet}
            all_header_align.append(row)

        if args.alignment_only:
            for row in compare_data_cells_alignment_and_format(
                ws1, ws2, one_cell_per_column=True
            ):
                row = {**row, "Sheet": sheet}
                all_align_and_format.append(row)

    print()
    print("=== Excel comparison ===")
    print(f"File 1: {os.path.basename(args.file1)}")
    print(f"File 2: {os.path.basename(args.file2)}")
    if args.alignment_only:
        print(
            "Mode: header alignment + one paired-data cell per column (alignment & format); "
            "empty paired columns → 'no data' (green). (--alignment-only)"
        )
    elif args.numeric_column_totals:
        print(
            "Quick numeric column totals: enabled (--numeric-column-totals); "
            "pandas sums only (no Excel format checks on columns)."
        )
    if args.row_counts:
        print("Per-sheet row counts: enabled (--row-counts; openpyxl max_row).")
    print()

    if args.row_counts:
        _print_section(
            "Per-sheet row counts (openpyxl Worksheet.max_row; Δ = File2 − File1 when both have the sheet) —",
            row_count_rows,
            "No worksheets.",
        )

    _print_section(
        "Sheet empty vs data mismatch — (same sheet name, one workbook has no cell data)",
        all_sheet_presence,
        "None — both files have data on each compared sheet, or both are empty.",
    )

    if not args.alignment_only:
        _print_section(
            "Column format summary (first data cell per column) —",
            all_column_fmt,
            "No differences.",
        )
        _print_section(
            f"Formatting cell mismatch — (~{args.sample_fraction:.0%} sample of rows with data in both files, per column)",
            all_cell_fmt,
            "No mismatches in sample.",
        )
        _print_section(
            "Amount total mismatch — (currency columns by Excel number format)",
            [r for r in all_currency_report if r.get("Totals_Match") == "No"],
            "No mismatches.",
        )
        if args.numeric_column_totals:
            _print_section(
                "Quick numeric column totals — (pandas only; no Excel format checks)",
                all_quick_numeric,
                "No shared columns with coercible numeric values (excluding datetime), or no data.",
            )

    _print_section(
        "Header alignment mismatch —",
        all_header_align,
        "No mismatches.",
    )

    if args.alignment_only:
        _print_section(
            "Data cell alignment & number format — (first row per column where both files have data; "
            "Note: no data = no such row, green in Excel/HTML)",
            all_align_and_format,
            "No issues — no mismatches on compared cells and no empty paired columns.",
        )
    else:
        _print_section(
            "Data alignment mismatch — (rows with data in both files)",
            all_data_align,
            "No mismatches.",
        )

    if args.output:
        write_report(
            args.output,
            formatting_cells_sampled=all_cell_fmt if not args.alignment_only else None,
            data_align=all_data_align if not args.alignment_only else None,
            header_align=all_header_align,
            currency_totals=all_currency_report if not args.alignment_only else None,
            numeric_column_totals=all_numeric_report if not args.alignment_only else None,
            quick_numeric_column_totals=all_quick_numeric
            if (not args.alignment_only and args.numeric_column_totals)
            else None,
            column_formatting=all_column_fmt if not args.alignment_only else None,
            sheet_presence=all_sheet_presence,
            data_align_with_format=all_align_and_format if args.alignment_only else None,
            row_counts=row_count_rows if args.row_counts else None,
        )
        print(f"Report written to: {args.output}")


if __name__ == "__main__":
    main()
