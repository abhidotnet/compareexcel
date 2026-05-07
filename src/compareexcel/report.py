"""Write comparison results to Excel or HTML."""

from __future__ import annotations

import html
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from compareexcel.core import ALIGN_ONLY_NO_DATA_NOTE

# Light yellow fill for alignment mismatch cells in Excel reports (RRGGBB).
_ALIGNMENT_MISMATCH_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")
_HTML_ALIGN_BG = "#fff9c4"
_EXCEL_GREEN_FONT = Font(color="006100")
_HTML_NO_DATA_COLOR = "#006100"

_HEADER_ALIGN_COLS = frozenset({"File1_Header_Alignment", "File2_Header_Alignment"})
_DATA_ALIGN_COLS = frozenset({"File1_Alignment", "File2_Alignment"})


def write_report(
    output_path: str | Path,
    *,
    formatting_cells_sampled,
    data_align,
    header_align,
    currency_totals,
    column_formatting: list | None = None,
    sheet_presence: list | None = None,
    data_align_with_format: list | None = None,
):
    """Write report; format is chosen from file suffix (.html vs .xlsx)."""
    path = Path(output_path)
    suffix = path.suffix.lower()
    if suffix in (".htm", ".html"):
        write_report_html(
            path,
            formatting_cells_sampled=formatting_cells_sampled,
            data_align=data_align,
            header_align=header_align,
            currency_totals=currency_totals,
            column_formatting=column_formatting,
            sheet_presence=sheet_presence,
            data_align_with_format=data_align_with_format,
        )
    else:
        write_report_excel(
            path,
            formatting_cells_sampled=formatting_cells_sampled,
            data_align=data_align,
            header_align=header_align,
            currency_totals=currency_totals,
            column_formatting=column_formatting,
            sheet_presence=sheet_presence,
            data_align_with_format=data_align_with_format,
        )


def write_report_excel(
    output_file: str | Path,
    *,
    formatting_cells_sampled,
    data_align,
    header_align,
    currency_totals,
    column_formatting: list | None = None,
    sheet_presence: list | None = None,
    data_align_with_format: list | None = None,
):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        wrote = False
        if sheet_presence:
            pd.DataFrame(sheet_presence).to_excel(
                writer, sheet_name="Sheet_Blank_Mismatch", index=False
            )
            wrote = True
        if column_formatting:
            pd.DataFrame(column_formatting).to_excel(
                writer, sheet_name="Column_Format_Summary", index=False
            )
            wrote = True
        if formatting_cells_sampled:
            pd.DataFrame(formatting_cells_sampled).to_excel(
                writer, sheet_name="Cell_Format_Sampled", index=False
            )
            wrote = True
        if currency_totals:
            pd.DataFrame(currency_totals).to_excel(
                writer, sheet_name="Amount_Total_Mismatch", index=False
            )
            wrote = True
        if header_align:
            pd.DataFrame(header_align).to_excel(
                writer, sheet_name="Header_Alignment_Mismatch", index=False
            )
            _after_alignment_sheet(writer, "Header_Alignment_Mismatch", _HEADER_ALIGN_COLS)
            wrote = True
        if data_align:
            pd.DataFrame(data_align).to_excel(
                writer, sheet_name="Data_Alignment_Mismatch", index=False
            )
            _after_alignment_sheet(writer, "Data_Alignment_Mismatch", _DATA_ALIGN_COLS)
            wrote = True
        if data_align_with_format:
            pd.DataFrame(data_align_with_format).to_excel(
                writer, sheet_name="Data_Cell_Align_Format", index=False
            )
            _style_data_cell_align_format_openpyxl(writer.book["Data_Cell_Align_Format"])
            wrote = True
        if not wrote:
            pd.DataFrame([{"Message": "No differences in this report."}]).to_excel(
                writer, sheet_name="Summary", index=False
            )


def _highlight_alignment_columns_openpyxl(ws, column_titles: frozenset[str]) -> None:
    """Apply yellow fill to listed header columns for all data rows (row 2+)."""
    if ws.max_row < 2:
        return
    headers = [c.value for c in ws[1]]
    col_indexes = [
        idx + 1
        for idx, h in enumerate(headers)
        if h is not None and str(h) in column_titles
    ]
    for c_idx in col_indexes:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=c_idx).fill = _ALIGNMENT_MISMATCH_FILL


def _after_alignment_sheet(writer: pd.ExcelWriter, sheet_name: str, columns: frozenset[str]) -> None:
    ws = writer.book[sheet_name]
    _highlight_alignment_columns_openpyxl(ws, columns)


def _style_data_cell_align_format_openpyxl(ws) -> None:
    """Green text for ``no data`` rows; yellow fill on alignment columns for mismatch rows."""
    if ws.max_row < 2:
        return
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]

    def col_idx(name: str) -> int | None:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    note_i = col_idx("Note")
    f1a = col_idx("File1_Alignment")
    f2a = col_idx("File2_Alignment")
    max_col = ws.max_column

    for r in range(2, ws.max_row + 1):
        note_val = ""
        if note_i:
            v = ws.cell(r, note_i).value
            if v is not None:
                note_val = str(v).strip().lower()
        if note_val == ALIGN_ONLY_NO_DATA_NOTE.lower():
            for c in range(1, max_col + 1):
                ws.cell(r, c).font = _EXCEL_GREEN_FONT
            continue
        for ci in (f1a, f2a):
            if ci:
                ws.cell(r, ci).fill = _ALIGNMENT_MISMATCH_FILL


def _df_section(
    title: str,
    rows: list,
    empty_message: str,
    *,
    highlight_alignment_columns: frozenset[str] | None = None,
) -> str:
    if not rows:
        return (
            f'<section class="block"><h2>{html.escape(title)}</h2>'
            f'<p class="ok">{html.escape(empty_message)}</p></section>'
        )
    df = pd.DataFrame(rows)
    subset = sorted(highlight_alignment_columns & set(df.columns)) if highlight_alignment_columns else []
    if subset:
        styler = df.style.set_properties(subset=subset, **{"background-color": _HTML_ALIGN_BG})
        styler = styler.hide(axis="index")
        table_html = styler.to_html(
            table_attributes='class="tbl" border="0"',
            escape=True,
        )
    else:
        table_html = df.to_html(classes="tbl", index=False, border=0, escape=True)
    return f'<section class="block"><h2>{html.escape(title)}</h2>{table_html}</section>'


def _df_section_data_cell_align_format(title: str, rows: list, empty_message: str) -> str:
    """HTML for Data_Cell_Align_Format: green ``no data`` rows, yellow alignment cells on mismatches."""
    if not rows:
        return (
            f'<section class="block"><h2>{html.escape(title)}</h2>'
            f'<p class="ok">{html.escape(empty_message)}</p></section>'
        )
    df = pd.DataFrame(rows)

    def _row_styles(row: pd.Series):
        note = str(row.get("Note", "")).strip().lower()
        if note == ALIGN_ONLY_NO_DATA_NOTE.lower():
            return [f"color: {_HTML_NO_DATA_COLOR}"] * len(row)
        r = row.get("Row")
        is_mismatch = bool(pd.notna(r) and str(r).strip() != "")
        return [
            f"background-color: {_HTML_ALIGN_BG}" if is_mismatch and col in _DATA_ALIGN_COLS else ""
            for col in row.index
        ]

    styler = df.style.apply(_row_styles, axis=1).hide(axis="index")
    table_html = styler.to_html(
        table_attributes='class="tbl" border="0"',
        escape=True,
    )
    return f'<section class="block"><h2>{html.escape(title)}</h2>{table_html}</section>'


def write_report_html(
    output_file: str | Path,
    *,
    formatting_cells_sampled,
    data_align,
    header_align,
    currency_totals,
    column_formatting: list | None = None,
    sheet_presence: list | None = None,
    data_align_with_format: list | None = None,
):
    parts = [
        "<!DOCTYPE html>",
        '<html lang="en"><head><meta charset="utf-8">',
        "<title>Excel comparison report</title>",
        "<style>",
        "body{font-family:system-ui,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:1.5rem;line-height:1.45;color:#1a1a1a;}",
        "h1{font-size:1.35rem;margin-bottom:0.5rem;}",
        "h2{font-size:1.05rem;margin:1.25rem 0 0.5rem;color:#333;border-bottom:1px solid #ddd;padding-bottom:0.25rem;}",
        ".block{margin-bottom:1.5rem;}",
        ".ok{color:#0d6b2c;}",
        ".tbl{border-collapse:collapse;width:100%;font-size:0.9rem;}",
        ".tbl th,.tbl td{border:1px solid #ccc;padding:6px 8px;text-align:left;}",
        ".tbl th{background:#f4f4f4;}",
        "</style></head><body>",
        "<h1>Excel comparison report</h1>",
    ]

    sp = sheet_presence or []
    parts.append(
        _df_section(
            "Sheet empty vs data mismatch (one workbook has no cell data)",
            sp,
            "None — both files have data on each compared sheet, or both are empty.",
        )
    )

    if column_formatting is not None:
        parts.append(
            _df_section(
                "Column format summary (first data cell)",
                column_formatting,
                "No column-level format differences.",
            )
        )
    if formatting_cells_sampled is not None:
        parts.append(
            _df_section(
                "Formatting cell mismatch — sampled cells (~10% of rows with data in both files, per column)",
                formatting_cells_sampled,
                "No sampled cell number_format mismatches.",
            )
        )
    if currency_totals is not None:
        parts.append(
            _df_section(
                "Amount total mismatch — currency columns",
                currency_totals,
                "No currency column total mismatches.",
            )
        )
    parts.append(
        _df_section(
            "Header alignment mismatch",
            header_align,
            "No header alignment mismatches.",
            highlight_alignment_columns=_HEADER_ALIGN_COLS,
        )
    )
    if data_align is not None:
        parts.append(
            _df_section(
                "Data alignment mismatch (rows with data in both files)",
                data_align,
                "No data alignment mismatches.",
                highlight_alignment_columns=_DATA_ALIGN_COLS,
            )
        )
    if data_align_with_format is not None:
        parts.append(
            _df_section_data_cell_align_format(
                "Data cell alignment & number format (one paired-data cell per column; "
                "green = no row with data in both files)",
                data_align_with_format,
                "No issues — no mismatches on compared cells and no empty paired columns.",
            )
        )
    parts.append("</body></html>")

    Path(output_file).write_text("\n".join(parts), encoding="utf-8")
